import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl
import zipfile
from UnprotectExcel_v1 import (
    unlock_workbook,
    unlock_worksheets,
    strip_protection_tags,
)


def test_unlock_workbook():
    wb = openpyxl.Workbook()
    wb.security.lockStructure = True
    unlock_workbook(wb)
    assert wb.security is None


def test_unlock_worksheets():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.protection.sheet = True
    ws.sheet_state = 'hidden'
    unlock_worksheets(wb)
    assert ws.sheet_state == 'visible'
    assert ws.protection is None


def test_strip_protection_tags(tmp_path: Path):
    path = tmp_path / "locked.xlsx"
    wb = openpyxl.Workbook()
    wb.security.lockStructure = True
    ws = wb.active
    ws.protection.sheet = True
    wb.save(path)
    strip_protection_tags(path)
    with zipfile.ZipFile(path) as z:
        data = z.read("xl/workbook.xml")
        assert b"workbookProtection" not in data
        sheet_data = z.read("xl/worksheets/sheet1.xml")
        assert b"sheetProtection" not in sheet_data

